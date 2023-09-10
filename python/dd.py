import requests



pgms = {
"p1a": """
# 1) AIM: Introduce the Python fundamentals, data types, operators, flow control andexception
# handling in Python
# a) Write a python program to find the best of two test average marks out of three test’smarks
# accepted from the user.
m1 = int(input("Enter the marks in the first test: "))
m2 = int(input("Enter the marks in the second test: "))
m3 = int(input("Enter the marks in the third test: "))
if m1 > m2:
        if m2 > m3:
                total = m1 + m2
        else:
                total = m1 + m3
elif m1 > m3:
        total = m1 + m2
else:
        total = m2 + m3
        avg = total / 2
print("The average of the best two test marks is:", avg)

# Output:
# CASE 1:
# Enter the marks in the first test: 20
# Enter the marks in second test: 15
# Enter the marks in third test: 22
# The average of the best two test marks is: 21.0

# CASE 2:
# Enter the marks in the first test: 24
# Enter the marks in second test: 25
# Enter the marks in third test: 23
# The average of the best two test marks is: 24.5
""",
"p1b": """
# b) Develop a Python program to check whether a given number is palindrome or not and also
# count the number of occurrences of each digit in the input number.
try:

        number = int(input("Enter the number: "))
        temp = number
        reverse = 0
        count = 0
        occurrences = {}
        while number > 0:
                digit = number % 10

                if digit in occurrences:
                         occurrences[digit] += 1
                else:
                        occurrences[digit] = 1

                reverse = reverse * 10 + digit
                number //= 10
                count += 1
        

        print("The reverse number is:", reverse)

        if temp == reverse:
                print("The number is a palindrome")
        else:
                print("The number is not a palindrome")

        print("Number of digits:", count)
        print("Digit occurrences:", occurrences)
except ValueError:
        print("Invalid input. Please enter a valid integer.")


# OUTPUT:
# Case 1:
# Enter the number: 1233
# The reverse number is: 3321
# The number is not a palindrome
# Number of digits: 4
# Digit occurrences: {3: 2, 2: 1, 1: 1}
# Case 2:
# Enter the number: 121
# The reverse number is: 121
# The number is a palindrome
# Number of digits: 3
# Digit occurrences: {1: 2, 2: 1}
# Case 3:
# Enter the number: ABC
# Invalid input. Please enter a valid integer.
""",
"p2a":"""
# 2) AIM: Demonstrating creation of functions, passing parameters and return values
# a) Defined a function F as Fn= Fn-1+Fn-2 . Write a python program which accepts a value for
# N (where N>0) as input and pass this value to the function. Display suitable error message if
# the condition for input value is not allowed.
def recur_fibo(n):

        if n <= 1:
                return n
        else:
                return(recur_fibo(n-1) + recur_fibo(n-2))
        
        
nterms = int(input("Enter the number"))

# check if the number of terms is valid
if nterms <= 0:
        
        print("Please enter a positive integer")
else:
        print("Fibonacci sequence:")
                
for i in range(nterms):
        print(recur_fibo(i))

# OUTPUT:
# CASE1:
# Enter the number5
# Fibonacci sequence:
# 0
# 1
# 1
# 2
# 3

# CASE 2:
# Enter the number-5
# Please enter a positive integer
""",
"p2b":"""
        # b) Develop a python program to convert binary to decimal, octal to hexadecimal using
# functions.
def binary_to_decimal(binary):
    decimal = 0
    power = 0
    while binary != 0:
        decimal += (binary % 10) * (2 ** power)
        binary //= 10
        power += 1
    return decimal

def octal_to_hexadecimal(octal):
    decimal = 0
    power = 0
    while octal != 0:
        decimal += (octal % 10) * (8 ** power)
        octal //= 10
        power += 1

    hexadecimal = ""
    while decimal != 0:
        remainder = decimal % 16
        if remainder < 10:
            hexadecimal = str(remainder) + hexadecimal
        else:
            hexadecimal = chr(ord('A') + remainder - 10) + hexadecimal
        decimal //= 16

    return hexadecimal

binary_number = input("Enter a binary number: ")
decimal_number = binary_to_decimal(int(binary_number))
print("Decimal equivalent:", decimal_number)

octal_number = input("Enter an octal number: ")
hexadecimal_number = octal_to_hexadecimal(int(octal_number))
print("Hexadecimal equivalent:", hexadecimal_number)



# Output:
# Enter a binary number: 101
# Decimal equivalent: 5
# Enter an octal number: 755
# Hexadecimal equivalent: 1ED
""",
"p3a":"""
# b) Develop a python program to convert binary to decimal, octal to hexadecimal using
# functions.
def binary_to_decimal(binary):
    decimal = 0
    power = 0
    while binary != 0:
        decimal += (binary % 10) * (2 ** power)
        binary //= 10
        power += 1
    return decimal

def octal_to_hexadecimal(octal):
    decimal = 0
    power = 0
    while octal != 0:
        decimal += (octal % 10) * (8 ** power)
        octal //= 10
        power += 1

    hexadecimal = ""
    while decimal != 0:
        remainder = decimal % 16
        if remainder < 10:
            hexadecimal = str(remainder) + hexadecimal
        else:
            hexadecimal = chr(ord('A') + remainder - 10) + hexadecimal
        decimal //= 16

    return hexadecimal

binary_number = input("Enter a binary number: ")
decimal_number = binary_to_decimal(int(binary_number))
print("Decimal equivalent:", decimal_number)

octal_number = input("Enter an octal number: ")
hexadecimal_number = octal_to_hexadecimal(int(octal_number))
print("Hexadecimal equivalent:", hexadecimal_number)



# Output:
# Enter a binary number: 101
# Decimal equivalent: 5
# Enter an octal number: 755
# Hexadecimal equivalent: 1ED
""",
"p3b":"""
# b) Write a Python program to find the string similarity between two given string.



import difflib
def string_similarity(str1, str2):
        result = difflib.SequenceMatcher(a=str1.lower(), b=str2.lower())
        return result.ratio()

str1 = input("Enter First string")
str2 = input("Enter Second string")
print("Original string:")
print(str1)
print(str2)
print("Similarity between two said strings:")
print(string_similarity(str1,str2))


# OUTPUT:
# Original string:
# Python Exercises
# Python Exercises
# Similarity between two said strings:
# 1.0
# Original string:
# Python Exercises
# Python Exercise
# Similarity between two said strings:
# 0.967741935483871
""",
"p4a":"""
# PYTHON LAB PROGRAM 4
# 4) AIM: Discuss different collections like list, tuple and dictionaries
# a) Write a python program to implement insertion sort and merge sort using lists

def insertion_sort(arr):
        for i in range(1, len(arr)):
                key = arr[i]
                j = i - 1
                while j >= 0 and key < arr[j]:
                        arr[j + 1] = arr[j]
                        j -= 1
                        arr[j + 1] = key

        return arr
#second method
def merge_sort(arr):
        if len(arr) > 1:
                mid = len(arr) // 2
                left_half = arr[:mid]
                right_half = arr[mid:]
                merge_sort(left_half)
                merge_sort(right_half)
                i = j = k = 0
                while i < len(left_half) and j < len(right_half):
                        if left_half[i] < right_half[j]:
                                        arr[k] = left_half[i]
                                        i += 1
                        else:
                                arr[k] = right_half[j]
                                j += 1
                        k += 1
                                
                while i < len(left_half):
                        arr[k] = left_half[i]
                        i += 1
                        k += 1

                while j < len(right_half):
                        arr[k] = right_half[j]
                        j += 1
                        k += 1
                return arr
my_list = [3, 1, 4, 1, 5, 9, 2, 6, 5, 3, 5]
sorted_list = insertion_sort(my_list)
print("Sorted list using Insertion sort:",sorted_list)
sorted_list = merge_sort(my_list)
print("Sorted list using Merge sort:",sorted_list)


# Output:
# Sorted list using Insertion sort: [1, 1, 2, 3, 3, 4, 5, 5, 5, 6, 9]
# Sorted list using Merge sort: [1, 1, 2, 3, 3, 4, 5, 5, 5, 6, 9]
""",
"p4b":"""
# b) Write a python program to convert Roman numbers into integer numbers usingdictionaries

class py_solution:
        def roman_to_int(self, s):
                rom_val = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100, 'D': 500, 'M': 1000}
                int_val = 0
                for i in range(len(s)):
                        if i > 0 and rom_val[s[i]] > rom_val[s[i - 1]]:
                                        int_val += rom_val[s[i]] - 2 * rom_val[s[i - 1]]
                        else:
                                int_val += rom_val[s[i]]
                return int_val
print(py_solution().roman_to_int('MMMCMLXXXVI'))
print(py_solution().roman_to_int('VIII'))
print(py_solution().roman_to_int('C'))

# OUTPUT:
# 3986
# 8
# 100
""",
"p5a":"""
# PYTHON LAB PROGRAM 5

# 5) AIM: Demonstration of pattern recognition with and without using regularexpressions.
# a) Write a function called isphonenumber () to recognize a pattern 415-555-242 without using regular expression
# and also write the code to recognize the same patternusing regular expression.
import re
def isphonenumber(numStr):
    
        if len(numStr) != 12:

                return False
        for i in range(len(numStr)):

                if i==3 or i==7:

                        if numStr[i] != "-":
                                return False
                else:
                                
                        if numStr[i].isdigit() == False:
                                        return False
        return True

def chkphonenumber(numStr):

        ph_no_pattern = re.compile(r'^\d{3}-\d{3}-\d{4}$')
        if ph_no_pattern.match(numStr):
                return True
        else:
                return False

ph_num = input("Enter a phone number : ")
print("Without using Regular Expression")
if isphonenumber(ph_num):
        print("Valid phone number")
else:
        print("Invalid phone number")

print("Using Regular Expression")
if chkphonenumber(ph_num):
        print("Valid phone number")
else:
        print("Invalid phone number")



# OUTPUT:
# CASE 1:
# Enter a phone number: 2e34-334-32
# Without using Regular Expression
# Invalid phone number
# Using Regular Expression
# Invalid phone number
# CASE 2:
# Enter a phone number: 234-567-2345
# Without using Regular Expression
# Valid phone number
# Using Regular Expression
# Valid phone number
""",
"p5b":"""
# b) Develop a python program that could search the text in a file for phone numbers(+919900889977)

# and email addresses (sample@gmail.com)

import re
phone_regex = re.compile(r'\+\d{12}')
email_regex = re.compile(r'[A-Za-z0-9._]+@[A-Za-z0-9]+\.[A-Z|a-z]{2,}')
with open('example.txt', 'r') as f:
        for line in f:
                matches = phone_regex.findall(line)
                for match in matches:
                        print(match)

                matches = email_regex.findall(line)
                for match in matches:
                        print(match)


# OUTPUT:
# tomcat@gmail.com
# jerrymouse@gmail.com
# spikethedog@gmail.com
# +919986059013
# +918310681568
""",
"p6a":"""
# PYTHON LAB PROGRAM 6
# 6) AIM: Demonstration Of reading, writing and organising files
# a) Write a python program to accept a file name from the user and perform the followingoperations
# 1. Display the first N line of the file
# 2. Find the frequency of occurrence of the word accepted from the user in the file
# Code:

import os.path
import sys

fname = input("Enter the filename: ")

if not os.path.isfile(fname):
        print("File", fname, "doesn't exist")
        sys.exit(0)

with open(fname, "r") as infile:
        lineList = infile.readlines()

for i in range(min(8, len(lineList))):
        print(i+1, ":", lineList[i].strip())

word = input("Enter a word: ")
cnt = 0
for line in lineList:
        cnt += line.count(word)

print("The word", word, "appears", cnt, "times in the file")

# OUTPUT:
# Enter the filename: example.txt
# 1 : Hello my name is Tom the cat.
# 2 : I like to play and work with my dear friend jerry mouse.
# 3 : We both have our office and email addresses
# 4 : They are tomcat@gmail.com, jerrymouse@gmail.com.
# 5 : Our friend spike has also joined us in our company.
# 6 : His email address is spikethedog@gmail.com.
# 7 : We all entertaint the children through our show.
# 8 : My phone number is +919986059013, +918310681568
# Enter a word: email
# The word email appears 2 times in the file
""",
"p6b":"""
# b) Write a python program to create a ZIP file of a particular folder which containsseveral files inside it.

import os
import sys
import pathlib
import zipfile
dirName = input("Enter Directory name that you want to backup : ")
if not os.path.isdir(dirName):
        print("Directory", dirName, "doesn't exists")
        sys.exit(0)
curDirectory = pathlib.Path(dirName)

with zipfile.ZipFile("myZip.zip", mode="w") as archive:
        for file_path in curDirectory.rglob("*"):
                archive.write(file_path, arcname=file_path.relative_to(curDirectory))
if os.path.isfile("myZip.zip"):
        print("Archive", "myZip.zip", "created successfully")
else:
        print("Error in creating zip archive")

# OUTPUT:
# Enter Directory name that you want to backup : New folder
# Archive myZip.zip created successfully
""",
"p7a":"""
# 7) AIM: Demonstration of concepts of classes, methods, objects and inheritance
# By using the concept of inheritance write a python program to find the area of triangle, circle and
# rectangle.
# Code:
import math
class Shape:
        def __init__(self):
                self.area = 0
                self.name = ""

        def showArea(self):
                print("The area of the", self.name, "is", self.area, "units")
class Circle(Shape):
        def __init__(self, radius):
                super().__init__()
                self.name = "Circle"
                self.radius = radius

        def calcArea(self):
                self.area = math.pi * self.radius * self.radius

class Rectangle(Shape):
        def __init__(self, length, breadth):
                super().__init__()
                self.name = "Rectangle"
                self.length = length
                self.breadth = breadth

        def calcArea(self):
                self.area = self.length * self.breadth

class Triangle(Shape):
        def __init__(self, base, height):
                super().__init__()
                self.name = "Triangle"
                self.base = base
                self.height = height

        def calcArea(self):
                self.area = self.base * self.height / 2

c1 = Circle(5)
c1.calcArea()
c1.showArea()
r1 = Rectangle(5, 4)
r1.calcArea()
r1.showArea()

t1 = Triangle(3, 4)
t1.calcArea()
t1.showArea()

# OUTPUT:
# The area of the Circle is 78.53981633974483 units
# The area of the Rectangle is 20 units
# The area of the Triangle is 6.0 units
""",
"p7b":"""
#pgm p7b
class Employee:
        def __init__(self):
                self.name = ""
                self.empId = ""
                self.dept = ""
                self.salary = 0

        def getEmpDetails(self):
                self.name = input("Enter Employee name: ")
                self.empId = input("Enter Employee ID: ")
                self.dept = input("Enter Employee Dept: ")
                self.salary = int(input("Enter Employee Salary: "))

        def showEmpDetails(self):
                print("Employee Details")
                print("Name:", self.name)
                print("ID:", self.empId)
                print("Dept:", self.dept)
                print("Salary:", self.salary)

        def updtSalary(self):
                self.salary = int(input("Enter new Salary: "))
                print("Updated Salary:", self.salary)

e1 = Employee()
e1.getEmpDetails()
e1.showEmpDetails()
e1.updtSalary()

# OUTPUT:
# Enter Employee name: Aaryan
# Enter Employee ID: 123
# Enter Employee Dept: cse
# Enter Employee Salary: 50000
# Employee Details
# Name: Aaryan
# ID: 123
# Dept: cse
# Salary: 50000
# Enter new Salary: 60600
# Updated Salary: 60600
""",
"p8a":"""
# 8) Aim: Demonstration of classes and methods with polymorphism and overriding
# a) Write a python program to find the whether the given input is palindrome or not
# (for both string and integer) using the concept of polymorphism and inheritance.
# Code:
class Palindrome:
        def __init__(self, data):
                self.data = data

        def is_palindrome(self):
                return str(self.data) == str(self.data)[::-1]

class StringPalindrome(Palindrome):
        def __init__(self, data):
                super().__init__(data)

        def is_palindrome(self):
                return super().is_palindrome()

class IntegerPalindrome(Palindrome):
        def __init__(self, data):
                super().__init__(data)

        def is_palindrome(self):
                return super().is_palindrome()

def check_palindrome(input_data):
        if isinstance(input_data, int):
                obj = IntegerPalindrome(input_data)
        elif isinstance(input_data, str):
                obj = StringPalindrome(input_data)
        else:
                raise ValueError("Input should be an integer or string")

        return obj.is_palindrome()

st = input("Enter a string : ")
stObj = StringPalindrome(st)
if stObj.is_palindrome():
        print("Given string is a Palindrome")
else:
        print("Given string is not a Palindrome")

val = int(input("Enter a integer : "))
intObj = IntegerPalindrome(val)
if intObj.is_palindrome():
        print("Given integer is a Palindrome")
else:
        print("Given integer is not a Palindrome")

# OUTPUT:
# Case1:
# Enter a string : AbA
# Given string is a Palindrome
# Enter a integer : 121
# Given integer is a Palindrome
# Case2:
# Enter a string : VTU
# Given string is not a Palindrome
# Enter a integer : 123
# Given integer is not a Palindrome
""",
"p9a":"""
# 9 Aim: Demonstration of working with excel spreadsheets and web scraping
# a) Write a python program to download the all XKCD comics
import requests
import os
def download_xkcd_comics():
        # Create a directory to store the comics
        os.makedirs('xkcd', exist_ok=True)
        # Start from the first comic
        url = 'https://xkcd.com/1/info.0.json'
        while True:
                # Fetch the comic metadata
                response = requests.get(url)
                response.raise_for_status()
                comic_data = response.json()
                # Extract the image URL
                image_url = comic_data['img']
                # Download the image
                response = requests.get(image_url)
                response.raise_for_status()
                # Save the image to the xkcd directory
                image_name = image_url.split('/')[-1]
                image_path = os.path.join('xkcd', image_name)
                with open(image_path, 'wb') as image_file:
                        image_file.write(response.content)
                # Print the comic title and number
                print(f"Downloaded: {comic_data['title']} - #{comic_data['num']}")
                # Check if there's a next comic
                if comic_data['num'] == 1:
                        break
                else:
                # Get the URL of the previous comic

                        url = f"https://xkcd.com/{comic_data['num'] - 1}/info.0.json"
                
print('All XKCD comics downloaded successfully!')
# Run the program
download_xkcd_comics()

# Output:
# Downloaded: Barrel - Part 1 - #1
# All XKCD comics downloaded successfully!
""",
"p9b":"""
# 9 b) Demonstrate python program to read the data from the spreadsheet and write the
# data in to the spreadsheet
# Code:
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook() 
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title = "Capital")
lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code =['KA', 'TS', 'TN']
sheet.cell(row = 1, column = 1).value = "State"
sheet.cell(row = 1, column = 2).value = "Language"
sheet.cell(row = 1, column = 3).value = "Code"
ft = Font(bold=True)
for row in sheet["A1:C1"]:
        for cell in row:
                cell.font = ft
for i in range(2,5):
        sheet.cell(row = i, column = 1).value = state[i-2]
        sheet.cell(row = i, column = 2).value = lang[i-2]
        sheet.cell(row = i, column = 3).value = code[i-2]
wb.save("demo.xlsx")
sheet = wb["Capital"]
sheet.cell(row = 1, column = 1).value = "State"
sheet.cell(row = 1, column = 2).value = "Capital"
sheet.cell(row = 1, column = 3).value = "Code"
ft = Font(bold=True)
for row in sheet["A1:C1"]:
        for cell in row:
                cell.font = ft
for i in range(2,5):
        sheet.cell(row = i, column = 1).value = state[i-2]
        sheet.cell(row = i, column = 2).value = capital[i-2]
        sheet.cell(row = i, column = 3).value = code[i-2]

wb.save("demo.xlsx")
srchCode = input("Enter state code for finding capital ")
for i in range(2,5):
        data = sheet.cell(row = i, column = 3).value
        if data == srchCode:
                print("Corresponding capital for code", srchCode, "is", sheet.cell(row = i, column =2).value)
sheet = wb["Language"]
srchCode = input("Enter state code for finding language ")
for i in range(2,5):
        data = sheet.cell(row = i, column = 3).value
        if data == srchCode:
                print("Corresponding language for code", srchCode, "is", sheet.cell(row = i, column =
2).value)
wb.close()


# Output:
# Enter state code for finding capital TS
# Corresponding capital for code TS is Hyderabad
# Enter state code for finding language TN
# Corresponding language for code TN is Tamil
""",
"p10a":"""
      # 10. AIM: Demonstration of working with PDF, word and JSON files
# a) Write a python program to combine select pages from many PDFsfrom PyPDF2 import
# PdfWriter, PdfReader
# Code:
from PyPDF2 import PdfWriter, PdfReader


def combine_pages(selected_pages, output_file):
    pdf_writer = PdfWriter()

    for pdf_file, page_numbers in selected_pages:
        pdf_reader = PdfReader(open(pdf_file, "rb"))
        for page_num in page_numbers:
            if 1 <= page_num <= len(pdf_reader.pages):
                page = pdf_reader.pages[page_num-1]
                pdf_writer.add_page(page)
            else:
                print(f"Page {page_num} is out of range for {pdf_file}.")

    with open(output_file, "wb") as output:
        pdf_writer.write(output)


# List of tuples containing PDF filenames and corresponding page numbers
selected_pages = [
    (
        "/Users/dhawanj/Desktop/engg_cutoff_gen_mock.pdf",
        [1, 2],
    ),  # Select page 1 and 2 from file1.pdf
    (
        
        "/Users/dhawanj/Desktop/engg_cutoff_gen_mock.pdf",
        [1, 2],
    ),  # Select page 2 and 4 from file2.pdf
    # Add more PDFs and page numbers as needed
]
output_filename = "combined_output.pdf"
combine_pages(selected_pages, output_filename)

""",
"p10b":"""
# b) Write a python program to fetch current weather data from the JSON file
# Code
import json
# Load the JSON data from file
with open('test/weather_data.json','r') as f:
        data = json.load(f)
# Extract the required weather data
current_temp = data['main']['temp']
humidity = data['main']['humidity']
weather_desc = data['weather'][0]['description']
# Display the weather data
print(f"Current temperature: {current_temp}°C")
print(f"Humidity: {humidity}%")
print(f"Weather description: {weather_desc}")

# Output:
# Current temperature: 15.45°C
# Humidity: 64%
# Weather description: clear sky
# this is json content copy this in another file and do it 
{
  "coord": {
    "lon": -73.99,
    "lat": 40.73
  },
  "weather": [
    {
      "id": 800,
      "main": "Clear",
      "description": "clear sky",
      "icon": "01d"
    }
  ],
  "base": "stations",
  "main": {
    "temp": 15.45,
    "feels_like": 12.74,
    "temp_min": 14.44,
    "temp_max": 16.11,
    "pressure": 1017,
    "humidity": 64
  },
  "visibility": 10000,
  "wind": {
    "speed": 4.63,

    "deg": 180
  },
  "clouds": {
    "all": 1
  },
  "dt": 1617979985,
  "sys": {
    "type": 1,
    "id": 5141,
    "country": "US",
    "sunrise": 1617951158,
    "sunset": 1618000213
  },
  "timezone": -14400,
  "id": 5128581,
  "name": "New York",
  "cod": 200
}
"""
}


def creatFile(dir, name, pgm):

    fullpath = dir+name
    f = open(fullpath, "w")
    f.write(pgms[pgm]);    
    


def sorry(pgm, loc="pyProject/"):
    ex = pgm+".py"
    try:
         pgms[pgm]
    except KeyError:
        print("given worng programe name --spell fuckker")  
        return   
   
    creatFile(loc, ex, pgm)
    print("fuck off")





def goc(pgm,loc="test/"):
        url = "https://raw.githubusercontent.com/dhawandj/cec/is/python/"  # Replace with the URL you want to make a request to
        nurl = pgm+".py"
        furl = url+nurl
        # Sending a GET request
        response = requests.get(furl)
        # Check the status code of the response
        if response.status_code == 200:
                print("Request was successful")
                print("Response content: Govindha Shiva ")
                # print(response.text)

                fullpath = loc+pgm+".py"
                f = open(fullpath, "w")
                f.write(response.text);    
                
        else:
                print("Request failed with status code:", response.status_code)

